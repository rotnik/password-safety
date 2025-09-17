import re
import hashlib
from collections import Counter
import matplotlib.pyplot as plt
import os
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

base_dir = os.path.dirname(os.path.abspath(__file__))
data_folder = os.path.join(base_dir, 'data')
results_folder = os.path.join(base_dir, 'results')

password_file = os.path.join(data_folder, 'password.txt')
passwords = []

try:
    with open(password_file, 'r', encoding='utf-8') as f:
        for line in f:
            pwd = line.strip()
            if pwd:
                passwords.append(pwd)
except FileNotFoundError:
    print(f"File {password_file} not found. Using sample passwords.")
    passwords = [
        "123456", "password", "qwerty", "letmein", "sunshine",
        "football", "iloveyou", "admin", "welcome", "monkey",
        "StrongPass1!", "My$ecureP@ss", "hello123", "abc123"
    ]

print(f"Loaded {len(passwords)} passwords.")

def check_password_strength(password):
    if len(password) < 8:
        return 'Weak'
    if re.fullmatch(r'[a-zA-Z]+', password):
        return 'Weak'
    if re.fullmatch(r'[a-zA-Z0-9]+', password):
        return 'Medium'
    if re.search(r'[^a-zA-Z0-9]', password):
        return 'Strong'
    return 'Medium'

def hash_password(password):
    return hashlib.sha256(password.encode()).hexdigest()

strengths = []
hashed_passwords = []

for pwd in passwords:
    strength = check_password_strength(pwd)
    hashed = hash_password(pwd)
    strengths.append(strength)
    hashed_passwords.append(hashed)

strength_count = Counter(strengths)

os.makedirs(results_folder, exist_ok=True)
output_file = os.path.join(results_folder, 'password_analysis_results.xlsx')

wb = Workbook()
ws = wb.active
ws.title = "Password Analysis"
ws.append(['Password', 'Strength', 'SHA-256 Hash'])

for pwd, strg, hsh in zip(passwords, strengths, hashed_passwords):
    ws.append([pwd, strg, hsh])

for col in ws.columns:
    max_length = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    ws.column_dimensions[col_letter].width = max_length + 2

wb.save(output_file)
print(f"Analysis exported to {output_file}")

labels = strength_count.keys()
counts = strength_count.values()

plt.figure(figsize=(6,4))
plt.bar(labels, counts, color=['red', 'orange', 'green'])
plt.title("Password Strength Distribution")
plt.xlabel("Strength")
plt.ylabel("Count")
plt.tight_layout()
plt.show()



